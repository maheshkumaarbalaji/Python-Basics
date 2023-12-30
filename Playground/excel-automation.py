import openpyxl
import random
import os
from datetime import datetime

Folder_Contents = os.listdir(os.getcwd())
if "Excel Documents" not in Folder_Contents:
    os.mkdir("Excel Documents")
    print("A new folder - 'Excel Documents' has been created in the current working directory.")
else:
    print("Folder - 'Excel Documents' already exists in the current working directory.")
Current_Datetime = datetime.now()
Timestamp = "{0}{1}{2}{3}{4}".format(str(Current_Datetime.year), str(Current_Datetime.month).rjust(2, '0'),
                                  str(Current_Datetime.day).rjust(2, '0'), str(Current_Datetime.hour).rjust(2, '0'), str(Current_Datetime.minute).rjust(2, '0'))
New_FileName = "Excel-" + Timestamp + ".xlsx"
print("New filename generated is:", New_FileName)
wb = openpyxl.Workbook()
ws = wb.active
for i in range(5):
    ws['A' + str(i + 1)] = random.randint(1, 100)
ws.append([670, 671, 672, 673, 674])
ws['B1'] = str(Current_Datetime)
wb.save("./Excel Documents/" + New_FileName)
